#!/usr/bin/env python3
"""
Download MyGuide-progress.xlsx from a SharePoint share link and regenerate
progress.json for the dashboard.

Reads the same two sheets the dashboard's "Import from Excel" button reads,
using the same coercion rules, so the automated output is byte-identical to
doing it by hand.

  Progress sheet  60 rows: Segment, Audience, Language, Chapter,
                           Progress %, Feedback 1, Feedback 2, Validated
  Slides sheet    one row per segment x audience x chapter
  Design sheet    one row per segment x audience: Segment, Audience, Status

Environment:
  XLSX_URL        SharePoint "Anyone with the link" URL   (required)
  OUT             output path       (default MyGuide-webflow/progress.json)
  ALLOW_PARTIAL   set to 1 to publish even when rows are missing
"""
import os, sys, json, re, io, datetime, urllib.request, urllib.error

OUT = os.environ.get('OUT', 'MyGuide-webflow/progress.json')
URL = os.environ.get('XLSX_URL', '').strip()
ALLOW_PARTIAL = os.environ.get('ALLOW_PARTIAL', '') == '1'


def die(msg, code=1):
    """Fail loudly and visibly. The ::error:: annotation puts the reason on the
    run's summary page, so nobody has to dig through step logs to find it."""
    print('ERROR: ' + msg, file=sys.stderr)
    one_line = ' '.join(msg.split())
    print('::error::' + one_line)
    summary = os.environ.get('GITHUB_STEP_SUMMARY')
    if summary:
        try:
            with open(summary, 'a', encoding='utf-8') as f:
                f.write('\n### Sync failed\n\n' + msg + '\n')
        except OSError:
            pass
    sys.exit(code)


def note(msg):
    print(msg)
    summary = os.environ.get('GITHUB_STEP_SUMMARY')
    if summary:
        try:
            with open(summary, 'a', encoding='utf-8') as f:
                f.write(msg + '\n\n')
        except OSError:
            pass


def _variants(url):
    """SharePoint share links come in several shapes and only some of them
    return the file itself rather than the Office web viewer. Try the known
    forms in order; the first one that yields a real .xlsx wins."""
    base = url.split('#')[0]
    seen, out = set(), []

    def add(u):
        if u and u not in seen:
            seen.add(u)
            out.append(u)

    # 1. the link as given, asking for the bytes rather than the viewer
    add(base + ('&' if '?' in base else '?') + 'download=1')
    # 2. same, with any tracking query dropped
    stem = base.split('?')[0]
    add(stem + '?download=1')
    # 3. the site's download.aspx, derived from a /:x:/s/<site>/<token> link
    m = re.match(r'^(https?://[^/]+)/:[a-z]:/[a-z]/([^/]+)/([^/?]+)', base, re.I)
    if m:
        host, site, token = m.group(1), m.group(2), m.group(3)
        add('%s/sites/%s/_layouts/15/download.aspx?share=%s' % (host, site, token))
        add('%s/teams/%s/_layouts/15/download.aspx?share=%s' % (host, site, token))
    return out


def download(url):
    attempts = []
    for candidate in _variants(url):
        req = urllib.request.Request(candidate, headers={
            'User-Agent': 'Mozilla/5.0 (compatible; MyGuide-dashboard-sync)',
            'Accept': '*/*'})
        try:
            with urllib.request.urlopen(req, timeout=120) as r:
                data = r.read()
                final = r.geturl()
        except urllib.error.HTTPError as e:
            attempts.append('%s -> HTTP %s' % (candidate, e.code))
            continue
        except Exception as e:
            attempts.append('%s -> %s' % (candidate, e))
            continue
        if data[:2] == b'PK':
            print('Downloaded the workbook (%d bytes) via: %s' % (len(data), candidate))
            return data
        snippet = data[:120].decode('utf-8', 'replace').replace('\n', ' ')
        signin = 'sign in' in snippet.lower() or 'Sign in' in data[:4000].decode('utf-8', 'replace')
        attempts.append('%s -> %s (%d bytes)%s'
                        % (candidate, 'HTML sign-in page' if signin else 'not a file',
                           len(data), '' if final == candidate else ' [redirected to %s]' % final))

    die('Could not download the workbook as a file. Every URL form was tried and '
        'none returned an .xlsx.\n\n'
        'Attempts:\n  ' + '\n  '.join(attempts) + '\n\n'
        'A sign-in page means the link is still scoped to your organisation. Fix it in '
        'three places, in this order:\n'
        '  1. SharePoint admin centre > Policies > Sharing > External sharing: set '
        'SharePoint to "Anyone". A site can never be more permissive than the tenant, '
        'so this has to be done first.\n'
        '  2. Sites > Active sites > your site > Settings > External sharing: also set '
        'to "Anyone".\n'
        '  3. Re-share the file: Share > click the audience button > "Anyone with the '
        'link" > Can view > Copy link. An existing organisation-scoped link does not '
        'become anonymous by itself.\n'
        'Then update the XLSX_URL secret with the NEW link and run the workflow again.')


def norm(v):
    return re.sub(r'[^a-z0-9]', '', str('' if v is None else v).lower())


def find_col(headers, *wanted):
    m = {norm(h): h for h in headers if h is not None}
    for w in wanted:
        n = norm(w)
        if n in m:
            return m[n]
        for k in m:
            if k.startswith(n):
                return m[k]
    return None


def to_pct(v):
    if v is None or v == '':
        return None
    if isinstance(v, bool):
        return None
    try:
        n = float(v) if not isinstance(v, str) else float(v.replace('%', '').replace(',', '.').strip())
    except ValueError:
        return None
    if 0 < n <= 1:
        n *= 100                      # a cell formatted as 80% arrives as 0.8
    n = round(n / 10.0) * 10
    return int(max(0, min(100, n)))


def to_fb(v):
    if v is None or v == '':
        return 0
    if isinstance(v, (int, float)) and not isinstance(v, bool):
        return int(v) if int(v) in (1, 2) else 0
    n = norm(v)
    if 'process' in n or n == 'done' or n == '2' or 'verwerkt' in n:
        return 2
    if 'review' in n or n == '1' or 'nazicht' in n or 'lopend' in n:
        return 1
    return 0


DP_STATES = ['not started', 'draft', 'submitted', 'approved']


def to_dp(v):
    if v is None or v == '':
        return 0
    if isinstance(v, (int, float)) and not isinstance(v, bool):
        n = int(v)
        return n if 0 <= n < len(DP_STATES) else 0
    n = norm(v)
    if 'approv' in n or 'goedgekeurd' in n or n == '3':
        return 3
    if 'submit' in n or 'ingediend' in n or n == '2':
        return 2
    if 'draft' in n or 'ontwerp' in n or n == '1':
        return 1
    return 0


def to_bool(v):
    if v is True:
        return True
    if isinstance(v, (int, float)) and not isinstance(v, bool):
        return int(v) == 1
    return norm(v) in ('yes', 'y', 'true', 'x', '1', 'ja', 'oui')


def sheet_rows(ws):
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return [], []
    headers = [('' if c is None else str(c).strip()) for c in rows[0]]
    out = []
    for r in rows[1:]:
        d = {}
        any_val = False
        for h, val in zip(headers, r):
            if not h:
                continue
            d[h] = '' if val is None else val
            if val not in (None, ''):
                any_val = True
        if any_val:
            out.append(d)
    return headers, out


def pick_sheet(wb, name, needed):
    for s in wb.sheetnames:
        if norm(s) == norm(name):
            return wb[s]
    for s in wb.sheetnames:
        headers, _ = sheet_rows(wb[s])
        if headers and all(find_col(headers, c) for c in needed):
            return wb[s]
    return None


def main():
    if not URL:
        die('XLSX_URL is not set. Add the SharePoint share link as a repository '
            'secret named XLSX_URL (Settings -> Secrets and variables -> Actions).')

    try:
        from openpyxl import load_workbook
    except ImportError:
        die('openpyxl is not installed.')

    # Existing progress.json is the template: it carries chapter titles, the
    # weighting settings and the deck structure, none of which come from Excel.
    if not os.path.exists(OUT):
        die('%s not found. It is the template for structure and settings; commit '
            'it once before enabling this workflow.' % OUT)
    with open(OUT, encoding='utf-8') as f:
        base = json.load(f)

    SEG   = base['segments']
    AUD   = base['audiences']
    LANG  = base['languages']
    NCH   = len(base['chapters'])
    expected_cells  = len(SEG) * len(AUD) * len(LANG) * NCH
    expected_slides = len(SEG) * len(AUD) * NCH

    wb = load_workbook(io.BytesIO(download(URL)), data_only=True, read_only=True)

    pws = pick_sheet(wb, 'Progress', ['segment', 'audience', 'language', 'chapter'])
    if pws is None:
        die('No Progress sheet found. Sheets in the workbook: %s' % ', '.join(wb.sheetnames))

    def match(v, options):
        n = norm(v)
        for o in options:
            if norm(o) == n:
                return o
        return None

    headers, rows = sheet_rows(pws)
    cSeg  = find_col(headers, 'segment')
    cAud  = find_col(headers, 'audience')
    cLang = find_col(headers, 'language')
    cCh   = find_col(headers, 'chapter')
    cP    = find_col(headers, 'progress', 'progress%', 'percent')
    cF1   = find_col(headers, 'feedback1', 'fb1')
    cF2   = find_col(headers, 'feedback2', 'fb2')
    cV    = find_col(headers, 'validated', 'validation')

    cells = {}
    unmatched = []
    for ix, row in enumerate(rows):
        seg  = match(row.get(cSeg), SEG)
        aud  = match(row.get(cAud), AUD)
        lang = match(row.get(cLang), LANG)
        try:
            ch = int(row.get(cCh))
        except (TypeError, ValueError):
            ch = 0
        if not (seg and aud and lang and 1 <= ch <= NCH):
            looks_like_data = str(row.get(cSeg, '')).strip() != '' and (
                str(row.get(cLang, '')).strip() != '' or str(row.get(cCh, '')).strip() != '')
            if looks_like_data and len(unmatched) < 10:
                unmatched.append('row %d: %s / %s / %s / %s' % (
                    ix + 2, row.get(cSeg), row.get(cAud), row.get(cLang), row.get(cCh)))
            continue
        key = '%s|%s|%s' % (aud, lang, seg)
        cells.setdefault(key, [None] * NCH)
        p = to_pct(row.get(cP)) if cP else None
        cells[key][ch - 1] = {
            'p':  0 if p is None else p,
            'f1': to_fb(row.get(cF1)) if cF1 else 0,
            'f2': to_fb(row.get(cF2)) if cF2 else 0,
            'v':  to_bool(row.get(cV)) if cV else False,
        }

    n_cells = sum(1 for v in cells.values() for c in v if c is not None)

    slides = {}
    n_slides = 0
    sws = pick_sheet(wb, 'Slides', ['segment', 'audience', 'chapter', 'slides'])
    if sws is not None:
        sheaders, srows = sheet_rows(sws)
        sSeg = find_col(sheaders, 'segment')
        sAud = find_col(sheaders, 'audience')
        sCh  = find_col(sheaders, 'chapter')
        sSl  = find_col(sheaders, 'slides', 'slide')
        for row in srows:
            seg = match(row.get(sSeg), SEG)
            aud = match(row.get(sAud), AUD)
            try:
                ch = int(row.get(sCh))
            except (TypeError, ValueError):
                ch = 0
            if not (seg and aud and 1 <= ch <= NCH):
                continue
            try:
                n = int(float(row.get(sSl) or 0))
            except (TypeError, ValueError):
                n = 0
            slides.setdefault('%s|%s' % (aud, seg), [0] * NCH)[ch - 1] = max(0, min(999, n))
            n_slides += 1

    design = {}
    n_design = 0
    expected_design = len(SEG) * len(AUD)
    dws = pick_sheet(wb, 'Design', ['segment', 'audience', 'status'])
    if dws is not None:
        dheaders, drows = sheet_rows(dws)
        dSeg = find_col(dheaders, 'segment')
        dAud = find_col(dheaders, 'audience')
        dSt  = find_col(dheaders, 'status', 'designproposal', 'design')
        for row in drows:
            seg = match(row.get(dSeg), SEG)
            aud = match(row.get(dAud), AUD)
            if not (seg and aud):
                continue
            design['%s|%s' % (aud, seg)] = to_dp(row.get(dSt)) if dSt else 0
            n_design += 1

    note('Matched %d/%d progress cells, %d/%d slide values and %d/%d design proposals.'
         % (n_cells, expected_cells, n_slides, expected_slides, n_design, expected_design))
    for u in unmatched:
        print('  unmatched %s' % u)

    if n_cells == 0:
        die('No row matched a known deck. Check that Segment reads %s, Audience reads %s '
            'and Language reads %s.' % ('/'.join(SEG), '/'.join(AUD), '/'.join(LANG)))
    if n_cells < expected_cells and not ALLOW_PARTIAL:
        die('Only %d of %d progress cells matched. Refusing to publish, because the '
            'missing chapters would be written as 0%%. Fix the workbook, or set '
            'ALLOW_PARTIAL=1 if this is deliberate.' % (n_cells, expected_cells))

    # fill gaps and write
    for seg in SEG:
        for aud in AUD:
            for lang in LANG:
                key = '%s|%s|%s' % (aud, lang, seg)
                cells.setdefault(key, [None] * NCH)
                for i in range(NCH):
                    if cells[key][i] is None:
                        cells[key][i] = {'p': 0, 'f1': 0, 'f2': 0, 'v': False}
            if n_slides:
                slides.setdefault('%s|%s' % (aud, seg), [0] * NCH)

    base['cells'] = cells
    if n_slides:
        base['slides'] = slides
    if n_design:
        for seg in SEG:
            for aud in AUD:
                design.setdefault('%s|%s' % (aud, seg), 0)
        base['design'] = design
    base['locked'] = True
    base['updated'] = datetime.date.today().isoformat()

    new = json.dumps(base, indent=2, ensure_ascii=False) + '\n'
    old = open(OUT, encoding='utf-8').read() if os.path.exists(OUT) else ''

    # Ignore a date-only difference so an unchanged workbook does not produce a
    # commit every fifteen minutes.
    def strip_date(s):
        return re.sub(r'"updated":\s*"[^"]*"', '"updated":""', s)

    if strip_date(new) == strip_date(old):
        note('No change in the figures — leaving progress.json alone.')
        return

    with open(OUT, 'w', encoding='utf-8') as f:
        f.write(new)
    note('Wrote %s' % OUT)


if __name__ == '__main__':
    main()
